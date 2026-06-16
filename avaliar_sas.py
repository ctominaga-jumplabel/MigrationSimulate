"""Avaliação determinística de complexidade dos arquivos .sas em ``all_sas/``.

Segue exatamente METODOLOGIA_AVALIACAO_SAS.md: 7 dimensões (D1-D7) com pesos
fixos (D1=0.15, D2=0.15, D3=0.15, D4=0.20, D5=0.10, D6=0.15, D7=0.10),
mapeamento score→categoria→horas com interpolação linear, multiplicadores
(piso 0.3, teto 2.0), detecção de duplicatas por nome-base, e xlsx de saída
com 3 abas (avaliacao, resumo_categoria, metodologia).
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import subprocess
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable

try:
    sys.stdout.reconfigure(encoding="utf-8")  # PowerShell costuma ser cp1252
except Exception:
    pass

# ---------------------------------------------------------------------------
# Bootstrap deps
# ---------------------------------------------------------------------------
try:
    import openpyxl  # noqa: F401
    import pandas as pd
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "pandas"])
    import openpyxl  # noqa: F401
    import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths / config
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parent
ALL_SAS = ROOT / "all_sas"
SAS_BY_EGP = ROOT / "sas_by_egp"
SOURCES = [ALL_SAS, SAS_BY_EGP]
METHOD_FILE = ROOT / "METODOLOGIA_AVALIACAO_SAS.md"
OUTPUT_XLSX = ROOT / "avaliacao_complexidade_sas.xlsx"

K_GLOBAL = 1.0  # calibração padrão

WEIGHTS = {
    "D1": 0.15,
    "D2": 0.15,
    "D3": 0.15,
    "D4": 0.20,
    "D5": 0.10,
    "D6": 0.15,
    "D7": 0.10,
}

CATEGORIES = [
    ("Trivial", 1.00, 1.79, 2.0, 6.0),
    ("Simples", 1.80, 2.59, 6.0, 16.0),
    ("Médio", 2.60, 3.39, 16.0, 40.0),
    ("Complexo", 3.40, 4.19, 40.0, 90.0),
    ("Muito Complexo", 4.20, 5.00, 90.0, 200.0),
]

# Lista exata da metodologia (§D3) - sem o '%'
MACRO_BUILTINS = {
    "let", "do", "if", "end", "mend", "put", "global", "local",
    "sysfunc", "scan", "eval", "sysevalf", "str", "nrstr",
    "bquote", "nrbquote", "superq", "symdel", "upcase", "lowcase",
    "substr", "index", "length", "trim", "left", "qsysfunc",
    "include", "macro",
}

# ---------------------------------------------------------------------------
# Leitura tolerante a encoding
# ---------------------------------------------------------------------------
def read_text(path: Path) -> tuple[str, str, bytes]:
    """Lê arquivo com fallback utf-8 → cp1252 → latin-1.
    Retorna (texto_decodificado, encoding_usado, bytes_brutos)."""
    raw = path.read_bytes()
    for enc in ("utf-8", "cp1252", "latin-1"):
        try:
            return raw.decode(enc), enc, raw
        except UnicodeDecodeError:
            continue
    return raw.decode("latin-1", errors="replace"), "latin-1?", raw


def hash_raw_bytes(raw: bytes) -> str:
    return hashlib.sha1(raw).hexdigest()


def normalize_for_hash(raw: bytes) -> bytes:
    """Normalização conservadora para o hash de duplicata:
    - remove BOM UTF-8
    - CRLF/CR → LF
    - rstrip por linha
    - remove linhas vazias no fim
    Não toca em comentários, strings ou whitespace interno.
    """
    b = raw
    if b.startswith(b"\xef\xbb\xbf"):
        b = b[3:]
    for enc in ("utf-8", "cp1252", "latin-1"):
        try:
            s = b.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        s = b.decode("latin-1", errors="replace")
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    lines = [ln.rstrip(" \t") for ln in s.split("\n")]
    while lines and lines[-1] == "":
        lines.pop()
    return "\n".join(lines).encode("utf-8")


def hash_normalized(raw: bytes) -> str:
    return hashlib.sha1(normalize_for_hash(raw)).hexdigest()


def egp_from_relpath(rel: str) -> str:
    """Retorna o nome do EGP quando o arquivo está em sas_by_egp/<egp>/...;
    vazio caso contrário."""
    parts = rel.split("/")
    if len(parts) >= 2 and parts[0] == "sas_by_egp":
        return parts[1]
    return ""


# ---------------------------------------------------------------------------
# Stripping: remove comentários e (opcionalmente) strings preservando
# quebras de linha. Estado-máquina simples cobrindo:
#   /* ... */ (bloco), * ... ;  (até próximo ;), %* ... ; (macro-comment),
#   "..." (aspas duplas), '...' (aspas simples), com escapes "" e ''.
# ---------------------------------------------------------------------------
def _strip(text: str, *, drop_strings: bool) -> str:
    out: list[str] = []
    i = 0
    n = len(text)
    state = "code"

    def back_significant_is(seps: tuple[str, ...]) -> bool:
        j = len(out) - 1
        while j >= 0 and out[j] in (" ", "\t"):
            j -= 1
        if j < 0:
            return True
        return out[j] in seps

    while i < n:
        c = text[i]
        if state == "code":
            # Block comment
            if c == "/" and i + 1 < n and text[i + 1] == "*":
                out.append("  ")
                i += 2
                state = "block"
                continue
            # Macro comment %* ... ;
            if c == "%" and i + 1 < n and text[i + 1] == "*":
                out.append("  ")
                i += 2
                state = "starcmt"
                continue
            # Star statement comment * ... ;  (somente em início de statement)
            if c == "*" and back_significant_is(("\n", ";")):
                out.append(" ")
                i += 1
                state = "starcmt"
                continue
            if c == '"':
                if drop_strings:
                    out.append(" ")
                else:
                    out.append('"')
                i += 1
                state = "dq"
                continue
            if c == "'":
                if drop_strings:
                    out.append(" ")
                else:
                    out.append("'")
                i += 1
                state = "sq"
                continue
            out.append(c)
            i += 1
        elif state == "block":
            if c == "*" and i + 1 < n and text[i + 1] == "/":
                out.append("  ")
                i += 2
                state = "code"
                continue
            out.append("\n" if c == "\n" else " ")
            i += 1
        elif state == "starcmt":
            if c == ";":
                out.append(";")
                i += 1
                state = "code"
                continue
            out.append("\n" if c == "\n" else " ")
            i += 1
        elif state == "dq":
            if c == '"':
                # escaped "" inside string
                if i + 1 < n and text[i + 1] == '"':
                    out.append("  " if drop_strings else '""')
                    i += 2
                    continue
                out.append(" " if drop_strings else '"')
                i += 1
                state = "code"
                continue
            if drop_strings:
                out.append("\n" if c == "\n" else " ")
            else:
                out.append(c)
            i += 1
        elif state == "sq":
            if c == "'":
                if i + 1 < n and text[i + 1] == "'":
                    out.append("  " if drop_strings else "''")
                    i += 2
                    continue
                out.append(" " if drop_strings else "'")
                i += 1
                state = "code"
                continue
            if drop_strings:
                out.append("\n" if c == "\n" else " ")
            else:
                out.append(c)
            i += 1
    return "".join(out)


def strip_comments(text: str) -> str:
    return _strip(text, drop_strings=False)


def strip_comments_and_strings(text: str) -> str:
    return _strip(text, drop_strings=True)


# ---------------------------------------------------------------------------
# Métricas brutas
# ---------------------------------------------------------------------------
PROC_NAME_RE = re.compile(r"(?im)^\s*proc\s+(\w+)")
DATA_STEP_RE = re.compile(r"(?im)^\s*data\s+[\w\.]+")
MACRO_CALL_RE = re.compile(r"%(\w+)\s*\(")
MACRO_DEF_RE = re.compile(r"(?i)%macro\s+(\w+)")
MACRO_TOKEN_RE = re.compile(r"(?i)%(macro|mend)\b")
JOIN_RE = re.compile(r"(?i)\b(?:inner|left|right|full|outer|cross)\s+join\b")
MERGE_RE = re.compile(r"(?im)^\s*merge\s+")
SET_MULTI_RE = re.compile(r"(?im)^\s*set\s+\S+\s+\S+")
TRANSPOSE_RE = re.compile(r"(?im)^\s*proc\s+transpose")
FIRST_LAST_RE = re.compile(r"(?i)\b(?:first|last)\.\w+\b")
RETAIN_RE = re.compile(r"(?i)\bretain\b")
LAG_RE = re.compile(r"(?i)\blag\d*\b")
DIF_RE = re.compile(r"(?i)\bdif\d*\b")
SUBQUERY_RE = re.compile(r"(?i)\(\s*select\b")
HAVING_RE = re.compile(r"(?i)\bhaving\b")
HASH_RE = re.compile(r"(?i)\bdeclare\s+hash\b")
ARRAY_RE = re.compile(r"(?im)^\s*array\s+\w+")
IF_THEN_RE = re.compile(r"(?im)(?<!%)\bif\s+.+\bthen\b")
SELECT_WHEN_RE = re.compile(r"(?is)\bselect\s*\(.*?\).*?\bwhen\b")
DO_INDEX_RE = re.compile(r"(?im)(?<!%)\bdo\s+\w+\s*=")
DO_WHILE_UNTIL_RE = re.compile(r"(?im)(?<!%)\bdo\s+(?:while|until)\b")
GOTO_LINK_RE = re.compile(r"(?im)\b(?:goto|link)\s+\w+")
LIBNAME_DB_RE = re.compile(
    r"(?im)^\s*libname\s+\w+\s+(?:oracle|sqlsvr|odbc|db2|teradata|postgres)\b"
)
LIBNAME_LOCAL_RE = re.compile(r"""(?im)^\s*libname\s+\w+\s+["']""")
IMPORT_EXPORT_RE = re.compile(r"(?im)^\s*proc\s+(?:import|export)\b")
INCLUDE_RE = re.compile(r"(?im)^\s*%include\s+")
FILENAME_EXT_RE = re.compile(r"(?i)\bfilename\s+\w+\s+(?:pipe|ftp|http|email)\b")
ODS_RE = re.compile(r"(?im)^\s*ods\s+\w+")
DDE_RE = re.compile(r"(?i)\bdde\b")
SOAP_HTTP_RE = re.compile(r"(?im)^\s*proc\s+(?:soap|http)\b")
IML_FCMP_OPT_RE = re.compile(r"(?im)^\s*proc\s+(?:iml|fcmp|optmodel)\b")
LET_RE = re.compile(r"(?im)^\s*%let\s+")
LIBNAME_ANY_RE = re.compile(r"(?im)^\s*libname\s+\w+")
FILENAME_ANY_RE = re.compile(r"(?im)^\s*filename\s+\w+")
MACRO_IF_RE = re.compile(r"(?i)%if\b")
MACRO_DO_RE = re.compile(r"(?i)%do\b")
BLOCK_OPENERS_RE = re.compile(r"(?i)(?<!%)\b(do|select|end)\b")
MAGIC_NUMBER_RE = re.compile(r"\b\d{5,}\b")
HARDCODED_DRV_RE = re.compile(r"""["'][A-Za-z]:[\\/]""")
HARDCODED_UNIX_RE = re.compile(r"""["']/[a-z]+/""")
NAME_LITERAL_RE = re.compile(r"""'[^'\n]{1,80}\s[^'\n]{0,80}'n""")
MOJIBAKE_RE = re.compile(r"Ã[ §£©¡³ªºµ\xa0­\"‡]|Ãƒ")

# Para o stack DO, ignorar tokens "end=" (option) e "end of file" não há;
# basta tokenizar palavras 'do', 'select', 'end' fora de strings/comentários.


def count_macro_calls(text: str) -> int:
    return sum(
        1 for m in MACRO_CALL_RE.finditer(text)
        if m.group(1).lower() not in MACRO_BUILTINS
    )


def macro_nesting_depth(text: str) -> int:
    depth = 0
    max_d = 0
    for m in MACRO_TOKEN_RE.finditer(text):
        if m.group(1).lower() == "macro":
            depth += 1
            if depth > max_d:
                max_d = depth
        else:
            depth = max(0, depth - 1)
    return max_d


def max_do_nesting(text: str) -> int:
    """Profundidade máxima de DO ignorando %do (macro). Considera SELECT...END
    para casar corretamente as marcas END, mas só conta DO no resultado."""
    stack: list[str] = []
    cur_do = 0
    max_do = 0
    for m in BLOCK_OPENERS_RE.finditer(text):
        tok = m.group(1).lower()
        if tok == "end":
            if stack:
                opener = stack.pop()
                if opener == "do":
                    cur_do = max(0, cur_do - 1)
        else:
            stack.append(tok)
            if tok == "do":
                cur_do += 1
                if cur_do > max_do:
                    max_do = cur_do
    return max_do


def proc_distribution(text: str) -> Counter:
    return Counter(m.group(1).lower() for m in PROC_NAME_RE.finditer(text))


def count_lines(raw: str, comments_only: str) -> tuple[int, int, int]:
    """Retorna (total, blank, comment_only)."""
    raw_lines = raw.splitlines() if raw else []
    com_lines = comments_only.splitlines() if comments_only else []
    total = len(raw_lines)
    # padronizar comprimento
    if len(com_lines) < total:
        com_lines = com_lines + [""] * (total - len(com_lines))
    blank = 0
    comment_only = 0
    for r, c in zip(raw_lines, com_lines):
        if r.strip() == "":
            blank += 1
        elif c.strip() == "":
            # tinha conteúdo mas sumiu após remoção de comentários
            comment_only += 1
    return total, blank, comment_only


# ---------------------------------------------------------------------------
# Scoring por dimensão
# ---------------------------------------------------------------------------
def score_D1(loc_eff: int) -> int:
    if loc_eff <= 50:
        return 1
    if loc_eff <= 200:
        return 2
    if loc_eff <= 600:
        return 3
    if loc_eff <= 1500:
        return 4
    return 5


def score_D2(n_data: int, n_proc: int) -> int:
    t = n_data + n_proc
    if t <= 3:
        return 1
    if t <= 10:
        return 2
    if t <= 25:
        return 3
    if t <= 60:
        return 4
    return 5


def score_D3(n_macros_def: int, n_macro_calls: int, has_if_and_do: bool,
             nest_depth: int) -> int:
    if n_macros_def > 15 or n_macro_calls > 80 or nest_depth >= 3:
        return 5
    if n_macros_def >= 7 or n_macro_calls >= 31 or nest_depth >= 2:
        return 4
    if n_macros_def >= 3 or n_macro_calls >= 11 or has_if_and_do:
        return 3
    if n_macros_def >= 1 or n_macro_calls >= 3:
        return 2
    return 1


def score_D4_points(metrics: dict) -> int:
    pts = 0
    pts += 2 * metrics["n_joins_sql"]
    pts += 2 * metrics["n_merges"]
    pts += 1 * metrics["n_set_multi"]
    pts += 3 * metrics["n_transpose"]
    pts += 1 if metrics["has_by_first_last"] else 0
    pts += metrics["n_retain_lag_dif_constructs"]
    pts += 2 * metrics["n_sql_subquery"]
    pts += 1 * metrics["n_having"]
    pts += 3 * (1 if metrics["has_hash_object"] else 0)
    pts += 1 * metrics["n_arrays"]
    return pts


def score_D4(metrics: dict) -> tuple[int, int]:
    pts = score_D4_points(metrics)
    if pts <= 2:
        return 1, pts
    if pts <= 6:
        return 2, pts
    if pts <= 15:
        return 3, pts
    if pts <= 30:
        return 4, pts
    return 5, pts


def score_D5(n_if_then: int, n_select_when: int, n_do_loops: int,
             has_goto: bool, max_do_nest: int) -> int:
    total = n_if_then + n_select_when + n_do_loops
    if total > 100 or max_do_nest >= 4 or has_goto:
        return 5
    if total >= 41:
        return 4
    if total >= 16:
        return 3
    if total >= 4:
        return 2
    return 1


def score_D6_points(metrics: dict) -> int:
    pts = 0
    pts += 3 * metrics["n_libname_db"]
    pts += 1 * metrics["n_libname_local"]
    pts += 2 * metrics["n_import_export"]
    pts += 2 * metrics["n_include"]
    pts += 4 * metrics["n_filename_external"]
    pts += 1 * metrics["n_ods"]
    pts += 5 * (1 if metrics["has_dde"] else 0)
    pts += 4 * metrics["n_soap_http"]
    return pts


def score_D6(metrics: dict) -> tuple[int, int]:
    pts = score_D6_points(metrics)
    if pts <= 1:
        return 1, pts
    if pts <= 4:
        return 2, pts
    if pts <= 10:
        return 3, pts
    if pts <= 20:
        return 4, pts
    return 5, pts


def score_D7(comment_ratio: float, n_hardcoded: int, n_magic: int,
             has_name_literal: bool) -> int:
    # cascata top-down: pior caso primeiro
    if has_name_literal and n_hardcoded > 3 and comment_ratio < 0.05:
        return 5
    if comment_ratio < 0.05 and n_hardcoded > 3:
        return 4
    if comment_ratio < 0.05 or (1 <= n_hardcoded <= 3) or n_magic > 5:
        return 3
    if comment_ratio < 0.15:
        return 2
    if comment_ratio >= 0.15 and n_hardcoded == 0 and n_magic == 0:
        return 1
    return 2


# ---------------------------------------------------------------------------
# Categoria e interpolação
# ---------------------------------------------------------------------------
def category_for(score: float) -> tuple[str, float, float, float, float]:
    for name, lo, hi, hmin, hmax in CATEGORIES:
        if lo <= score <= hi:
            return name, lo, hi, hmin, hmax
    if score < CATEGORIES[0][1]:
        return CATEGORIES[0]
    return CATEGORIES[-1]


def interpolate_hours(score: float) -> tuple[str, float, float, float]:
    name, lo, hi, hmin, hmax = category_for(score)
    span = hi - lo
    pos = 0.0 if span == 0 else (score - lo) / span
    pos = min(1.0, max(0.0, pos))
    hours = hmin + pos * (hmax - hmin)
    return name, hmin, hmax, hours


# ---------------------------------------------------------------------------
# Duplicatas
# ---------------------------------------------------------------------------
DUP_SUFFIX_PATTERNS = [
    # numeric/copy markers
    r"\s*\(\d+\)\s*$",
    r"\s*-\s*[Cc]opia\d*\s*$",
    r"\s*_[Cc]opia\d*\s*$",
    r"\s*-\s*[Cc]opy\d*\s*$",
    r"\s*_[Cc]opy\d*\s*$",
    r"(?i)\s*_bkp\d*\s*$",
    r"(?i)\s*_bk\d*\s*$",
    r"(?i)\s*_backup\d*\s*$",
    r"(?i)\s*_v\d+\s*$",
    r"(?i)\s*_ver\d+\s*$",
    r"(?i)\s*_old\d*\s*$",
    r"(?i)\s*_new\d*\s*$",
    r"(?i)\s*_final\d*\s*$",
    r"(?i)\s*_tmp\d*\s*$",
    # date/version markers (sufixos típicos de cópia versionada por data)
    r"_\d{8}$",                            # _20200818
    r"_\d{6}$",                            # _200818 / _311020
    r"_\d{1,2}_\d{1,2}_\d{2,4}$",          # _08_03_2021, _8_3_21
    r"_\d{1,2}_\d{1,2}$",                  # _31_10
    # trailing separators
    r"[\s_\-]+$",
]


def name_base(filename: str) -> str:
    stem = os.path.splitext(filename)[0]
    changed = True
    while changed:
        changed = False
        for p in DUP_SUFFIX_PATTERNS:
            new = re.sub(p, "", stem)
            if new != stem:
                stem = new
                changed = True
    return stem.strip().lower()


# ---------------------------------------------------------------------------
# Avaliação por arquivo
# ---------------------------------------------------------------------------
def evaluate_file(path: Path, rel: str) -> dict:
    raw, encoding_used, raw_bytes = read_text(path)
    h_raw = hash_raw_bytes(raw_bytes)
    h_norm = hash_normalized(raw_bytes)
    egp = egp_from_relpath(rel)
    comments_only_stripped = strip_comments(raw)         # mantém strings
    cleaned = strip_comments_and_strings(raw)            # remove comentários + strings

    # Linhas
    loc_total, loc_blank, loc_comment = count_lines(raw, comments_only_stripped)
    loc_effective = loc_total - loc_blank - loc_comment
    comment_ratio = (loc_comment / loc_effective) if loc_effective > 0 else 0.0

    # Métricas brutas (no texto limpo)
    proc_dist = proc_distribution(cleaned)
    n_proc_steps = sum(proc_dist.values())
    n_data_steps = len(DATA_STEP_RE.findall(cleaned))

    n_macros_def = len(MACRO_DEF_RE.findall(cleaned))
    n_macro_calls = count_macro_calls(cleaned)
    macro_nest = macro_nesting_depth(cleaned)
    has_macro_if_and_do = (
        MACRO_IF_RE.search(cleaned) is not None
        and MACRO_DO_RE.search(cleaned) is not None
    )

    n_joins_sql = len(JOIN_RE.findall(cleaned))
    n_merges = len(MERGE_RE.findall(cleaned))
    n_set_multi = len(SET_MULTI_RE.findall(cleaned))
    n_transpose = len(TRANSPOSE_RE.findall(cleaned))
    has_by_first_last = FIRST_LAST_RE.search(cleaned) is not None
    n_retain_lag_dif = sum([
        1 if RETAIN_RE.search(cleaned) else 0,
        1 if LAG_RE.search(cleaned) else 0,
        1 if DIF_RE.search(cleaned) else 0,
    ])
    n_sql_subquery = len(SUBQUERY_RE.findall(cleaned))
    n_having = len(HAVING_RE.findall(cleaned))
    has_hash_object = HASH_RE.search(cleaned) is not None
    n_arrays = len(ARRAY_RE.findall(cleaned))
    has_arrays = n_arrays > 0

    n_if_then = len(IF_THEN_RE.findall(cleaned))
    n_select_when = len(SELECT_WHEN_RE.findall(cleaned))
    n_do_loops = (
        len(DO_INDEX_RE.findall(cleaned))
        + len(DO_WHILE_UNTIL_RE.findall(cleaned))
    )
    max_do_nest = max_do_nesting(cleaned)
    has_goto_link = GOTO_LINK_RE.search(cleaned) is not None

    n_libname_db = len(LIBNAME_DB_RE.findall(cleaned))
    n_libname_local = len(LIBNAME_LOCAL_RE.findall(cleaned))
    n_import_export = len(IMPORT_EXPORT_RE.findall(cleaned))
    n_include = len(INCLUDE_RE.findall(cleaned))
    n_filename_external = len(FILENAME_EXT_RE.findall(cleaned))
    n_ods = len(ODS_RE.findall(cleaned))
    has_dde = DDE_RE.search(cleaned) is not None
    n_soap_http = len(SOAP_HTTP_RE.findall(cleaned))
    has_iml_fcmp_optmodel = IML_FCMP_OPT_RE.search(cleaned) is not None

    # Métricas em RAW (precisam de strings):
    n_hardcoded_paths = (
        len(HARDCODED_DRV_RE.findall(raw))
        + len(HARDCODED_UNIX_RE.findall(raw))
    )
    has_name_literal = NAME_LITERAL_RE.search(raw) is not None
    n_magic_numbers = len(MAGIC_NUMBER_RE.findall(cleaned))
    has_mojibake = bool(MOJIBAKE_RE.search(raw))

    # Suporte a params-only / classificação
    n_let = len(LET_RE.findall(cleaned))
    n_libname_total = len(LIBNAME_ANY_RE.findall(cleaned))
    n_filename_total = len(FILENAME_ANY_RE.findall(cleaned))
    is_params_only = (
        n_data_steps == 0
        and n_proc_steps == 0
        and n_macros_def == 0
        and (n_let + n_libname_total + n_filename_total) > 0
    )

    # Scores por dimensão
    d1 = score_D1(loc_effective)
    d2 = score_D2(n_data_steps, n_proc_steps)
    d3 = score_D3(n_macros_def, n_macro_calls, has_macro_if_and_do, macro_nest)
    metrics_d4 = {
        "n_joins_sql": n_joins_sql,
        "n_merges": n_merges,
        "n_set_multi": n_set_multi,
        "n_transpose": n_transpose,
        "has_by_first_last": has_by_first_last,
        "n_retain_lag_dif_constructs": n_retain_lag_dif,
        "n_sql_subquery": n_sql_subquery,
        "n_having": n_having,
        "has_hash_object": has_hash_object,
        "n_arrays": n_arrays,
    }
    d4, d4_pts = score_D4(metrics_d4)
    d5 = score_D5(n_if_then, n_select_when, n_do_loops, has_goto_link, max_do_nest)
    metrics_d6 = {
        "n_libname_db": n_libname_db,
        "n_libname_local": n_libname_local,
        "n_import_export": n_import_export,
        "n_include": n_include,
        "n_filename_external": n_filename_external,
        "n_ods": n_ods,
        "has_dde": has_dde,
        "n_soap_http": n_soap_http,
    }
    d6, d6_pts = score_D6(metrics_d6)
    d7 = score_D7(comment_ratio, n_hardcoded_paths, n_magic_numbers, has_name_literal)

    score_pond = (
        d1 * WEIGHTS["D1"] + d2 * WEIGHTS["D2"] + d3 * WEIGHTS["D3"]
        + d4 * WEIGHTS["D4"] + d5 * WEIGHTS["D5"] + d6 * WEIGHTS["D6"]
        + d7 * WEIGHTS["D7"]
    )
    score_pond = round(score_pond, 4)

    categoria, h_min, h_max, h_interp = interpolate_hours(score_pond)

    # Observações
    obs: list[str] = []
    if has_mojibake or (encoding_used != "utf-8" and any(ord(ch) > 127 for ch in raw[:5000])):
        obs.append("encoding suspeito")
    if has_dde:
        obs.append("DDE")
    if has_iml_fcmp_optmodel:
        obs.append("IML/FCMP/OPTMODEL")
    if macro_nest >= 3:
        obs.append(f"macro depth={macro_nest}")
    if is_params_only:
        obs.append("apenas parâmetros")

    return {
        "file_name": path.name,
        "file_path": rel,
        "egp": egp,
        "size_kb": round(path.stat().st_size / 1024, 2),
        "content_hash_raw": h_raw,
        "content_hash_normalized": h_norm,
        "loc_total": loc_total,
        "loc_effective": loc_effective,
        "loc_comment": loc_comment,
        "comment_ratio": round(comment_ratio, 4),
        "n_data_steps": n_data_steps,
        "n_proc_steps": n_proc_steps,
        "proc_distribution": json.dumps(dict(proc_dist), ensure_ascii=False),
        "n_macros_defined": n_macros_def,
        "n_macro_calls": n_macro_calls,
        "macro_nesting_depth": macro_nest,
        "n_joins_sql": n_joins_sql,
        "n_merges": n_merges,
        "n_transpose": n_transpose,
        "has_hash_object": has_hash_object,
        "has_arrays": has_arrays,
        "n_if_then": n_if_then,
        "n_select_when": n_select_when,
        "n_do_loops": n_do_loops,
        "max_do_nesting": max_do_nest,
        "has_goto_link": has_goto_link,
        "n_libname_db": n_libname_db,
        "n_libname_local": n_libname_local,
        "n_import_export": n_import_export,
        "n_include": n_include,
        "has_dde": has_dde,
        "has_iml_fcmp_optmodel": has_iml_fcmp_optmodel,
        "n_hardcoded_paths": n_hardcoded_paths,
        "score_D1_tamanho": d1,
        "score_D2_estruturas": d2,
        "score_D3_macro": d3,
        "score_D4_manipulacao": d4,
        "score_D5_logica": d5,
        "score_D6_integracoes": d6,
        "score_D7_qualidade": d7,
        "score_ponderado": score_pond,
        "categoria": categoria,
        "horas_base_min": h_min,
        "horas_base_max": h_max,
        "horas_base_interpoladas": round(h_interp, 2),
        "multiplicadores_aplicados": "",
        "multiplicador_total": 1.0,
        "horas_estimadas": 0.0,
        "duplicate_group_id": "",
        "duplicate_detection_method": "none",
        "is_likely_duplicate": False,
        "canonical_file": "",
        "same_namebase_group_size": 1,
        "same_namebase_unique_normalized_hashes": 1,
        "is_params_only": is_params_only,
        "observacoes": "; ".join(obs),
        # campos auxiliares (não exportados):
        "_encoding_used": encoding_used,
        "_d4_points": d4_pts,
        "_d6_points": d6_pts,
        "_has_mojibake": has_mojibake,
        "_macro_depth": macro_nest,
    }


# ---------------------------------------------------------------------------
# Multiplicadores e duplicatas (pós-processamento)
# ---------------------------------------------------------------------------
def apply_multipliers_and_dups(rows: list[dict]) -> None:
    """Detecção de duplicatas por conteúdo (hash normalizado) e aplicação
    dos multiplicadores. Nome-base só serve para colunas de auditoria; não
    influencia o multiplicador de duplicata."""

    # ---- 1) agrupar por hash normalizado ----
    by_norm: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        by_norm[r["content_hash_normalized"]].append(r)

    # ---- 2) colunas de auditoria por nome-base ----
    by_namebase: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        by_namebase[name_base(r["file_name"])].append(r)
    for nb, items in by_namebase.items():
        size = len(items)
        unique_norm = len({r["content_hash_normalized"] for r in items})
        for r in items:
            r["same_namebase_group_size"] = size
            r["same_namebase_unique_normalized_hashes"] = unique_norm

    # ---- 3) duplicatas reais (grupos com >=2 arquivos pelo hash normalizado) ----
    for norm_hash, items in by_norm.items():
        if len(items) <= 1:
            # único: não-duplicata
            r = items[0]
            r["duplicate_group_id"] = ""
            r["duplicate_detection_method"] = "none"
            r["is_likely_duplicate"] = False
            r["canonical_file"] = ""
            continue
        # escolher canônico: maior loc_effective; tie maior size_kb; tie menor file_path
        ordered = sorted(
            items,
            key=lambda r: (-r["loc_effective"], -r["size_kb"], r["file_path"]),
        )
        canon = ordered[0]
        all_same_raw = len({r["content_hash_raw"] for r in items}) == 1
        method = "exact_content" if all_same_raw else "normalized_content"
        for r in items:
            r["duplicate_group_id"] = norm_hash
            r["duplicate_detection_method"] = method
            if r is canon:
                r["is_likely_duplicate"] = False
                r["canonical_file"] = ""
            else:
                r["is_likely_duplicate"] = True
                r["canonical_file"] = canon["file_path"]

    # ---- 4) multiplicadores ----
    for r in rows:
        applied: list[str] = []
        mult = 1.0
        if r["is_params_only"]:
            mult *= 0.4
            applied.append("params_only(x0.4)")
        if r["has_dde"]:
            mult *= 1.4
            applied.append("DDE(x1.4)")
        if r["has_iml_fcmp_optmodel"]:
            mult *= 1.5
            applied.append("IML/FCMP/OPTMODEL(x1.5)")
        if r["_has_mojibake"]:
            mult *= 1.1
            applied.append("encoding(x1.1)")
        if r["_macro_depth"] >= 3:
            mult *= 1.2
            applied.append("macro_depth>=3(x1.2)")
        if r["is_likely_duplicate"]:
            mult *= 0.3
            applied.append("duplicata(x0.3)")
        mult = max(0.3, min(2.0, mult))
        r["multiplicadores_aplicados"] = "; ".join(applied)
        r["multiplicador_total"] = round(mult, 4)
        r["horas_estimadas"] = round(
            r["horas_base_interpoladas"] * mult * K_GLOBAL, 2
        )


# ---------------------------------------------------------------------------
# Saída XLSX
# ---------------------------------------------------------------------------
COLUMN_ORDER = [
    "file_name", "file_path", "egp", "size_kb",
    "content_hash_raw", "content_hash_normalized",
    "loc_total", "loc_effective", "loc_comment", "comment_ratio",
    "n_data_steps", "n_proc_steps", "proc_distribution",
    "n_macros_defined", "n_macro_calls", "macro_nesting_depth",
    "n_joins_sql", "n_merges", "n_transpose",
    "has_hash_object", "has_arrays", "n_if_then", "n_select_when",
    "n_do_loops", "max_do_nesting", "has_goto_link", "n_libname_db",
    "n_libname_local", "n_import_export", "n_include", "has_dde",
    "has_iml_fcmp_optmodel", "n_hardcoded_paths",
    "score_D1_tamanho", "score_D2_estruturas", "score_D3_macro",
    "score_D4_manipulacao", "score_D5_logica", "score_D6_integracoes",
    "score_D7_qualidade", "score_ponderado", "categoria",
    "horas_base_min", "horas_base_max", "horas_base_interpoladas",
    "multiplicadores_aplicados", "multiplicador_total", "horas_estimadas",
    "duplicate_group_id", "duplicate_detection_method",
    "is_likely_duplicate", "canonical_file",
    "same_namebase_group_size", "same_namebase_unique_normalized_hashes",
    "is_params_only", "observacoes",
]

HOUR_COLS = {
    "horas_base_min", "horas_base_max", "horas_base_interpoladas",
    "horas_estimadas",
}

# colunas textuais que precisam ser escapadas contra interpretação como fórmula
SAFE_STR_COLS = {
    "file_name", "file_path", "egp", "canonical_file",
    "proc_distribution", "multiplicadores_aplicados", "observacoes",
    "content_hash_raw", "content_hash_normalized", "duplicate_group_id",
    "duplicate_detection_method", "categoria",
}


def safe_excel_cell(value, col_name: str):
    """Evita que o Excel interprete strings começando com =,+,-,@ como fórmula."""
    if col_name in SAFE_STR_COLS and isinstance(value, str) and value:
        if value[0] in ("=", "+", "-", "@"):
            return "'" + value
    return value


def write_xlsx(rows: list[dict], out: Path, methodology_text: str) -> None:
    wb = Workbook()
    # --- aba 1: avaliacao ---
    ws = wb.active
    ws.title = "avaliacao"
    ws.append(COLUMN_ORDER)
    for r in rows:
        ws.append([safe_excel_cell(r.get(c, ""), c) for c in COLUMN_ORDER])
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    # formato 1 casa decimal para colunas de horas
    for col_idx, name in enumerate(COLUMN_ORDER, start=1):
        if name in HOUR_COLS:
            letter = get_column_letter(col_idx)
            for cell in ws[letter][1:]:
                cell.number_format = "0.0"
    # largura aproximada
    for col_idx, name in enumerate(COLUMN_ORDER, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = max(12, min(40, len(name) + 2))

    # --- aba 2: resumo_categoria ---
    ws2 = wb.create_sheet("resumo_categoria")
    df = pd.DataFrame(rows)
    summary = (
        df.groupby("categoria")
        .agg(
            n_arquivos=("file_name", "count"),
            horas_estimadas_total=("horas_estimadas", "sum"),
            horas_medias=("horas_estimadas", "mean"),
        )
        .reset_index()
    )
    total_h = summary["horas_estimadas_total"].sum()
    summary["pct_do_total_horas"] = (
        summary["horas_estimadas_total"] / total_h * 100
        if total_h > 0 else 0
    )
    # reordenar segundo a ordem das categorias
    order = {name: i for i, (name, *_rest) in enumerate(CATEGORIES)}
    summary["__o"] = summary["categoria"].map(order)
    summary = summary.sort_values("__o").drop(columns="__o").reset_index(drop=True)

    cols2 = ["categoria", "n_arquivos", "horas_estimadas_total", "horas_medias", "pct_do_total_horas"]
    ws2.append(cols2)
    for _, row in summary.iterrows():
        ws2.append([row[c] for c in cols2])
    for cell in ws2[1]:
        cell.font = bold
    ws2.freeze_panes = "A2"
    for col_idx, name in enumerate(cols2, start=1):
        letter = get_column_letter(col_idx)
        ws2.column_dimensions[letter].width = 22
        if name in {"horas_estimadas_total", "horas_medias", "pct_do_total_horas"}:
            for cell in ws2[letter][1:]:
                cell.number_format = "0.0"

    # linha de total
    ws2.append([
        "TOTAL",
        int(summary["n_arquivos"].sum()),
        float(summary["horas_estimadas_total"].sum()),
        float(summary["horas_estimadas_total"].sum() / summary["n_arquivos"].sum())
            if summary["n_arquivos"].sum() else 0.0,
        100.0,
    ])
    last_row = ws2.max_row
    for c in ws2[last_row]:
        c.font = bold

    # --- aba 3: metodologia ---
    ws3 = wb.create_sheet("metodologia")
    ws3.append(["texto"])
    for cell in ws3[1]:
        cell.font = bold
    for line in methodology_text.splitlines() or [""]:
        ws3.append([line])
    ws3.freeze_panes = "A2"
    ws3.column_dimensions["A"].width = 120
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws3.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    wb.save(out)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> int:
    if not METHOD_FILE.exists():
        print(f"ERRO: metodologia não encontrada: {METHOD_FILE}", file=sys.stderr)
        return 1

    files: list[Path] = []
    for src in SOURCES:
        if not src.exists():
            print(f"AVISO: diretório ausente, ignorado: {src}", file=sys.stderr)
            continue
        found = sorted(src.rglob("*.sas"))
        print(f"Encontrados {len(found)} arquivos .sas em {src}")
        files.extend(found)
    print(f"Total a processar: {len(files)}")

    rows: list[dict] = []
    read_failures: list[tuple[str, str]] = []
    for i, p in enumerate(files, start=1):
        rel = p.relative_to(ROOT).as_posix()
        try:
            row = evaluate_file(p, rel)
            rows.append(row)
        except Exception as exc:  # noqa: BLE001
            read_failures.append((rel, repr(exc)))
        if i % 2500 == 0:
            print(f"  ... {i}/{len(files)} processados")

    apply_multipliers_and_dups(rows)

    methodology_text = METHOD_FILE.read_text(encoding="utf-8", errors="replace")
    write_xlsx(rows, OUTPUT_XLSX, methodology_text)
    print(f"\nGerado: {OUTPUT_XLSX}")

    # --- resumo console ---
    df = pd.DataFrame(rows)
    cat_order = [c[0] for c in CATEGORIES]
    print("\n=== Distribuição por categoria ===")
    by_cat = (
        df.groupby("categoria")
        .agg(
            n=("file_name", "count"),
            horas=("horas_estimadas", "sum"),
            media=("horas_estimadas", "mean"),
        )
        .reindex(cat_order)
        .fillna(0)
    )
    print(by_cat.to_string())

    total_horas = float(df["horas_estimadas"].sum())
    print(f"\nTotal de horas estimadas: {total_horas:,.1f} h  (K={K_GLOBAL})")
    n_dup = int(df["is_likely_duplicate"].sum())
    print(f"Duplicatas detectadas: {n_dup}")

    print(f"\nFalhas de leitura: {len(read_failures)}")
    for f, err in read_failures[:20]:
        print(f"  - {f}: {err}")

    print("\n=== Top 10 por horas estimadas ===")
    top10 = df.nlargest(10, "horas_estimadas")[
        ["file_name", "file_path", "score_ponderado", "categoria",
         "horas_estimadas", "multiplicadores_aplicados"]
    ]
    print(top10.to_string(index=False))

    # ---------- VALIDAÇÕES ----------
    print("\n" + "=" * 60)
    print("VALIDAÇÕES")
    print("=" * 60)

    # (1) totais
    print(f"\n[1] Total de arquivos processados: {len(df)}")

    # (2) por origem
    df["__source"] = df["file_path"].str.split("/").str[0]
    print("\n[2] Total por origem:")
    print(df["__source"].value_counts().to_string())

    # (3) duplicatas por método
    grp_raw = df.groupby("content_hash_raw").size()
    excess_raw = int((grp_raw - 1).clip(lower=0).sum())
    grp_norm = df.groupby("content_hash_normalized").size()
    excess_norm = int((grp_norm - 1).clip(lower=0).sum())
    marcadas = int(df["is_likely_duplicate"].sum())
    print(f"\n[3] Duplicatas:")
    print(f"     - excesso por content_hash_raw       : {excess_raw}")
    print(f"     - excesso por content_hash_normalized: {excess_norm}")
    print(f"     - marcadas na planilha               : {marcadas}")

    # (4) quantidade de grupos de duplicatas
    n_groups = int((grp_norm > 1).sum())
    print(f"\n[4] Grupos de duplicatas (hash normalizado, >=2 arquivos): {n_groups}")

    # (5) top 10 maiores grupos
    print("\n[5] Top 10 maiores grupos de duplicatas (por content_hash_normalized):")
    top_groups = grp_norm.sort_values(ascending=False).head(10)
    for h, sz in top_groups.items():
        members = df[df["content_hash_normalized"] == h]
        # canônico = is_likely_duplicate == False
        canon_row = members[~members["is_likely_duplicate"]].iloc[0]
        others = (
            members[members["is_likely_duplicate"]]["file_path"]
            .head(3).tolist()
        )
        print(f"  - tamanho={sz}  canônico={canon_row['file_path']}")
        for o in others:
            print(f"      ex: {o}")

    # (6) grupos por nome-base com mais de um hash normalizado
    df["__nb"] = df["file_name"].apply(name_base)
    nb_stats = df.groupby("__nb").agg(
        n=("file_name", "size"),
        uniq_hash=("content_hash_normalized", "nunique"),
    )
    nb_mixed = nb_stats[nb_stats["uniq_hash"] > 1]
    print(f"\n[6] Grupos por nome-base com >1 hash normalizado: {len(nb_mixed)}")
    print("    Top 5 (auditoria: novo critério não colapsa cegamente):")
    for nb, row in nb_mixed.sort_values("n", ascending=False).head(5).iterrows():
        print(f"  - nome-base='{nb}'  n_files={row['n']}  n_hashes_distintos={row['uniq_hash']}")

    # (7) re-abrir o xlsx e validar escape em colunas textuais
    print("\n[7] Validação de escape de fórmula (relendo o xlsx):")
    text_cols_check = [
        "file_name", "file_path", "egp", "canonical_file",
        "proc_distribution", "multiplicadores_aplicados", "observacoes",
    ]
    from openpyxl import load_workbook  # local import
    wb_check = load_workbook(OUTPUT_XLSX, read_only=True, data_only=False)
    ws_check = wb_check["avaliacao"]
    header_row = next(ws_check.iter_rows(min_row=1, max_row=1, values_only=True))
    idx_map = {n: i for i, n in enumerate(header_row)}
    check_idx = [(c, idx_map[c]) for c in text_cols_check if c in idx_map]
    escaped_ok = 0
    not_escaped = []
    for row in ws_check.iter_rows(min_row=2, values_only=True):
        for col_name, i in check_idx:
            v = row[i]
            if isinstance(v, str) and v:
                first = v[0]
                if first in ("=", "+", "-", "@"):
                    not_escaped.append((col_name, v[:80]))
                elif first == "'" and len(v) > 1 and v[1] in ("=", "+", "-", "@"):
                    escaped_ok += 1
    wb_check.close()
    print(f"     - células corretamente escapadas com apóstrofo: {escaped_ok}")
    print(f"     - células ainda iniciadas com =/+/-/@ (deveriam ser 0): {len(not_escaped)}")
    for c, v in not_escaped[:5]:
        print(f"       {c}: {v!r}")

    # (8) file_path 100% preenchido
    empty_fp = int(df["file_path"].fillna("").eq("").sum())
    print(f"\n[8] file_path vazio: {empty_fp} de {len(df)}")

    # (9) recalcula score_ponderado e horas_estimadas e confirma bate
    rec_score = (
        df["score_D1_tamanho"] * WEIGHTS["D1"]
        + df["score_D2_estruturas"] * WEIGHTS["D2"]
        + df["score_D3_macro"] * WEIGHTS["D3"]
        + df["score_D4_manipulacao"] * WEIGHTS["D4"]
        + df["score_D5_logica"] * WEIGHTS["D5"]
        + df["score_D6_integracoes"] * WEIGHTS["D6"]
        + df["score_D7_qualidade"] * WEIGHTS["D7"]
    ).round(4)
    diff_score = (rec_score - df["score_ponderado"]).abs().max()
    rec_horas = (
        df["horas_base_interpoladas"] * df["multiplicador_total"] * K_GLOBAL
    ).round(2)
    diff_horas = (rec_horas - df["horas_estimadas"]).abs().max()
    print(f"\n[9] Reconciliação:")
    print(f"     - max abs diff score_ponderado: {diff_score}")
    print(f"     - max abs diff horas_estimadas: {diff_horas}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
